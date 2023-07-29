import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color, Fill
from openpyxl import cell

template_file = "template.xlsx" 
new_file = "test.xlsx"

book = load_workbook(template_file)
writer = pd.ExcelWriter(new_file, engine = 'openpyxl') 
writer.book = book
ws = book.worksheets("Data")

## ws is a openpypxl worksheet object
_cell = ws.cell()

# Font properties
_cell.style.font.color.index = Color.BLACK
_cell.style.font.name = 'Arial'
_cell.style.font.size = 8
_cell.style.font.bold = True

pd.read_csv('adjust.csv').to_excel(writer, sheet_name = 'Data', index = False)

# Saving the file as 'test.xlsx'
writer.save()
writer.close()