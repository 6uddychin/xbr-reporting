import pandas as pd
from datetime import datetime, date, timedelta
import sys
import openpyxl as xl
from openpyxl import load_workbook
import time



activeDate = pd.to_datetime(datetime.today() + timedelta(days=10))
reportWeek = (datetime.today() - timedelta(days=10)).strftime("%V")
reportName = str("WK" + reportWeek + "PartnerOps_WBR.xlsx")
tempReport = str("wk" + reportWeek + "temp.xlsx")
dfName = ("wk" + reportWeek + "_qbr.xlsx")



df1 = pd.read_csv('qbr.csv')
df1.to_excel(tempReport)

time.sleep(2)

df = pd.read_excel(tempReport)

# LIST OF FIELDS TO CHECK FOR CORRECT COLUMN
survey_request_columns_to_check = ["Survey Request DateTime", "Date site was turned over from BD to Ops"]
survey_complete_columns_to_check = ["Actual Survey Date", "On-site consultation actual date", "Survey Uploaded DateTime"]
asset_load_columns_to_check = ["Asset Load date","Date Trouble Ticket Created"]
installation_columns_columns_to_check = ["Activation Date","Installation Date"]
removal_columns_to_check = ["Removal Date","Scheduled Decommission date/time"]
active_pipeline_columns_to_check = ["Date/Time Closed", "Asset Load date","Date Trouble Ticket Created", "Activation Date","Installation Date"]

# GET THE FIRST NOT NULL VALUE IN THE FIELDS ABOVE
df['SurveyRequest'] = df[survey_request_columns_to_check].bfill(axis=1).iloc[:, 0]
df['SurveyComplete'] = df[survey_complete_columns_to_check ].bfill(axis=1).iloc[:, 0]
df['AssetLoad'] = df[asset_load_columns_to_check].bfill(axis=1).iloc[:, 0]
df['Activation'] = df[installation_columns_columns_to_check].bfill(axis=1).iloc[:, 0]
df['Decomm'] = df[removal_columns_to_check].bfill(axis=1).iloc[:, 0]

# GROUP THE COUNTRIES INTO REGIONS
def assign_region(country):
    if country in ["CA", "MX","US"]:
        return "NA"
    elif country in ["FR", "DE", "GB", "PT", "ES", "AT","IT"]:
        return "EU"
    elif country in ["JP","AU"]:
        return "APAC"
    else:
        return None
df['region'] = df['Country Code'].apply(assign_region)

# DEFINE THE PROGRAM
def assign_program(recordType):
    if recordType in ["Locker Onboarding", "Locker Onboarding NA","Odin Onboarding EU"]:
        return "Locker"
    elif recordType in ["Location Onboarding", "Location"]:
        return "Apt Locker Pro"
    elif recordType in ["Dobby Onboarding", "Dobby Onboarding EU", "Dobby Onboarding NA"]:
        return "Apt Locker"
    else:
        return None
df['program'] = df['Case Record Type'].apply(assign_program)

# CONVERT TIME/DATE FIELDS TO THE CORRECT DATE FORMATS FOR CYCLE TIMES
df['SurveyRequest'] = pd.to_datetime(df['SurveyRequest'])
df['SurveyComplete'] = pd.to_datetime(df['SurveyComplete'])
df['AssetLoad'] = pd.to_datetime(df['AssetLoad'])
df['Activation'] = pd.to_datetime(df['Activation'])

def check_InstallCycleTime (row):
    if row['Activation'] > row['AssetLoad']: 
        return (row['Activation'] - row['AssetLoad']).days
    else:
        return ""

def check_e2e (row):
    if row['Activation'] > row['SurveyRequest']: 
        return (row['Activation'] - row['SurveyRequest']).days
    else:
        return ""
    
def check_wastedTrip(row):
    if row['Survey Review outcome'] == "Wasted Trip":
        return row['SurveyComplete']
    else:
        return ""

def check_SurveyCycleTime (row):
    if row['SurveyComplete'] > row['SurveyRequest']:
        return (row['SurveyComplete'] - row['SurveyRequest']).days
    else:
        return ""
    
def activeFunc (x):
    if x == "NaN":
        activeDate = date.today() + timedelta(days=10)
        return activeDate
    else:
        df[active_pipeline_columns_to_check].bfill(axis=1).iloc[:, 0]


# RUNS THE FUNCTIONS AND ADDS THE COLUMNS TO THE DATAFRAME
df['region'] = df['Country Code'].apply(assign_region)
df['SurveyCycleTime'] = df.apply(check_SurveyCycleTime, axis=1)
df['InstallCycleTime'] = df.apply(check_InstallCycleTime, axis=1)
df['e2eCycleTime'] = df.apply(check_e2e, axis=1)
df['WatedTrip'] = df.apply(check_wastedTrip, axis = 1)
df['Active'] = df[active_pipeline_columns_to_check].bfill(axis=1).iloc[:, 0]
df['ActivePipeline'] = df['Active'].mask(df['Active'].isna(), activeDate)


def badSurvey(date1,date2):
    if date1 > date2:
        return False
    else:
        return True

df['BadSurveyData'] = df.apply(lambda row: badSurvey(row['SurveyComplete'], row['SurveyRequest']), axis = 1 )
df['BadInstall'] = df.apply(lambda row: badSurvey(row['Activation'], row['AssetLoad']), axis = 1 )
df['BadEnd2End'] = df.apply(lambda row: badSurvey(row['Activation'], row['SurveyRequest']), axis = 1 )


# SAVING AS XLSX FILE GETS AROUND HAVING TO SET 'STYLES' WHEN COPY/PASTING TO XLSX FORM CSV -WHICH CAUSES XLSX TO REPAIR THE WORKBOOK
dataFile = "transform.xlsx"

df.to_excel(
    tempReport, 
    header=[
    'region', 'Country Code','program','Status','SurveyRequest', 'SurveyComplete','WatedTrip',  'AssetLoad', 'Activation','Decomm' ,'Failed Install Date', 'ActivePipeline', 'CaseHold','SurveyCycleTime' ,'InstallCycleTime','e2eCycleTime','BadSurveyData','BadInstall','BadEnd2End'
    ], 
    columns=[
        'region', 'Country Code','program','Status','SurveyRequest', 'SurveyComplete','WatedTrip',  'AssetLoad', 'Activation','Decomm' , 'Failed Install Date','ActivePipeline','CaseHold', 'SurveyCycleTime' ,'InstallCycleTime','e2eCycleTime','BadSurveyData','BadInstall','BadEnd2End'
        ], encoding='utf-8', index=False)

# MAKE SURE THE FILE SAVES
time.sleep(30)

# QBR TEMPLATE
qbrFile = "qbr_template.xlsx"

# COMPILE THE FILE
wb1 = xl.load_workbook(tempReport)
ws1 = wb1.worksheets[0]
wb2 = xl.load_workbook(qbrFile)
wb2.active = wb2["Data"]
ws2 = wb2.active

# ADD ROW COUNTS AND COLUMNS COUNTS FOR COPY/PASTE LOOP
mr = ws1.max_row
mc = ws1.max_column

# MAKE SURE BOTH FILES LOADED. THIS MAY NEED TO BE ADJUSTED
time.sleep(30)

# COPY/PASTE LOOP
for i in range (1, mr +1):
    for j in range (1, mc + 1):
        c = ws1.cell(row= i, column = j)
        ws2.cell(row = i, column = j).value = c.value


# SAVES THE FINAL FILE WITH DATA TAB
wb2.save(reportName)

time.sleep(10)

copyValues = pd.read_excel(reportName, sheet_name="Working")
time.sleep(20)
copyValues = copyValues.drop(copyValues.columns[0], axis = 1)

output_file2 =  "qbr_format_test2.xlsx"
copyValues.to_excel(output_file2, sheet_name='table')

def copy_worksheet(source_ws, target_ws):
    for row in source_ws.iter_rows():
        for cell in row:
            target_ws[cell.coordinate].font = cell.font.copy()
            target_ws[cell.coordinate].border = cell.border.copy()
            target_ws[cell.coordinate].fill = cell.fill.copy()
            target_ws[cell.coordinate].number_format = cell.number_format
            target_ws[cell.coordinate].alignment = cell.alignment.copy()

def main():
    source_file = qbrFile
    source_sheet_name = "Working"
    output_file = "qbr_Final.xlsx"

    source_wb = xl.load_workbook(source_file)
    source_ws = source_wb[source_sheet_name]

    output_wb = xl.load_workbook(reportName)
    output_ws = output_wb.active
    output_ws.title = source_sheet_name

    copy_worksheet(source_ws, output_ws)
    output_wb.save(output_file)

if __name__ == "__main__":
    main()


