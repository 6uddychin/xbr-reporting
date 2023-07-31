import pandas as pd
import openpyxl as xl
from openpyxl import load_workbook
import datetime


wb1 = xl.load_workbook("WK29PartnerOps.xlsx")
mainTable = wb1["Working"]

# create new file
newBook = xl.Workbook()
newBook.save("testworkbook.xlsx")
newSheet = newBook["Sheet"]



mr = mainTable.max_row
mc = mainTable.max_column
print(mr)

# transfer info from one sheet to another
for i in range (1 , mr +1 ):
    for j in range (1, mc +1 ):
        c = mainTable.cell(row = i, column = j).value
        newSheet.cell(row = i, column = j) = c.value

newBook.save("testworkbook.xlsx")
