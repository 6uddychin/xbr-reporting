import pandas as pd
import openpyxl as xl
from openpyxl import load_workbook

# df = pd.read_csv(r'style_column.csv')
# readXLSX = pd.read_excel(r'style_test.xlsx')
# df.to_excel(r'style_test.xlsx', sheet_name='new', index=False, header=True)
#

# readXLSX = 'style_column.xlsx'
# df.to_excel(readXLSX, index=False)

dataFile = "transform.xlsx"
qbrFile = "qbr_test.xlsx"

# open workbooks
wb1 = xl.load_workbook(dataFile)
ws1 = wb1.worksheets[0]
wb2 = xl.load_workbook(qbrFile)
wb2.active = wb2["Data"]
ws2 = wb2.active

mr = ws2.max_row
mc = ws2.max_column

# loop to copy data
for i in range (1, mr +1):
    for j in range (1, mc + 1):
        c = ws1.cell(row= i, column = j)
        ws2.cell(row = i, column = j).value = c.value

wb2.save("test_loop.xlsx")

print(mr)