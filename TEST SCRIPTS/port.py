import openpyxl

def copy_worksheet(source_ws, target_ws):
    for row in source_ws.iter_rows():
        for cell in row:
            target_ws[cell.coordinate].value = cell.value
            target_ws[cell.coordinate].font = cell.font.copy()
            target_ws[cell.coordinate].border = cell.border.copy()
            target_ws[cell.coordinate].fill = cell.fill.copy()
            target_ws[cell.coordinate].number_format = cell.number_format
            target_ws[cell.coordinate].protection = cell.protection.copy()
            target_ws[cell.coordinate].alignment = cell.alignment.copy()
            target_ws[cell.coordinate].comment = cell.comment

def main():
    source_file = "WK29PartnerOps.xlsx"
    source_sheet_name = "Working"
    output_file = "testworkbook.xlsx"

    source_wb = openpyxl.load_workbook(source_file)
    source_ws = source_wb[source_sheet_name]

    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active
    output_ws.title = source_sheet_name

    copy_worksheet(source_ws, output_ws)
    output_wb.save(output_file)

if __name__ == "__main__":
    main()