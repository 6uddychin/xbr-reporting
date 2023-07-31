import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font

testCSV="style_columns.csv"
testXLSX="style.xlsx"

book = load_workbook(testXLSX)

#writer = pd.ExcelWriter(testCSV, engine = 'openpyxl') 
writer = book
ws = book.active
total_rows = ws.max_row

# opened, provisioned
print(book.sheetnames)
print(ws.max_row)

